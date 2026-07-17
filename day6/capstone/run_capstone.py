#!/usr/bin/env python3
"""
================================================================
Capstone: Python × MiniZinc × Excel × Matplotlib
กรณีศึกษา: Job Shop Scheduling สำหรับโรงงานผลิตโลหะ
================================================================

การทำงาน:
  1. อ่านข้อมูล orders จาก capstone_data.xlsx  (pandas + openpyxl)
  2. ส่ง parameters ให้ MiniZinc solver          (minizinc Python API)
  3. รับ schedule ที่ optimal                    (Gecode)
  4. แสดง Gantt chart ด้วย matplotlib

ติดตั้ง dependencies:
  pip install minizinc pandas openpyxl matplotlib

รัน:
  python run_capstone.py
================================================================
"""

import os
import sys

# ── ตรวจสอบ libraries ────────────────────────────────────────────
missing = []
for lib in ['minizinc', 'pandas', 'openpyxl', 'matplotlib']:
    try:
        __import__(lib)
    except ImportError:
        missing.append(lib)
if missing:
    print(f"กรุณาติดตั้ง: pip install {' '.join(missing)}")
    sys.exit(1)

import minizinc
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(HERE, 'capstone_data.xlsx')
MODEL_FILE = os.path.join(HERE, 'jobshop_model.mzn')
OUTPUT_IMG = os.path.join(HERE, 'schedule_result.png')


# ════════════════════════════════════════════════════════════════
# STEP 0: สร้างไฟล์ Excel ตัวอย่าง (ถ้ายังไม่มี)
# ════════════════════════════════════════════════════════════════

def create_sample_excel():
    """สร้าง capstone_data.xlsx สำหรับโรงงานโลหะ 4 orders × 3 tasks × 3 machines"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: orders ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'orders'
    header = ['JobID', 'JobName', 'Product', 'Quantity', 'Deadline (min)']
    ws.append(header)
    ws.append([1, 'OrderA', 'เก้าอี้เหล็ก',  50, 30])
    ws.append([2, 'OrderB', 'โต๊ะเหล็ก',      30, 35])
    ws.append([3, 'OrderC', 'ชั้นวางของ',      40, 28])
    ws.append([4, 'OrderD', 'โครงโครงสร้าง',  25, 40])

    # ── Sheet 2: machines ────────────────────────────────────────
    ws2 = wb.create_sheet('machines')
    ws2.append(['MachineID', 'MachineName', 'Specialty'])
    ws2.append([1, 'M_A', 'ตัด — Cutting (เร็วที่สุด แต่เชื่อมไม่ได้)'])
    ws2.append([2, 'M_B', 'เชื่อม — Welding (เร็วที่สุด แต่ตัดไม่ได้)'])
    ws2.append([3, 'M_C', 'All-rounder (ทุก task ได้ แต่ช้ากว่า)'])

    # ── Sheet 3: proc_times ──────────────────────────────────────
    # dur[job, task, machine] — 0 หมายถึงทำไม่ได้
    ws3 = wb.create_sheet('proc_times')
    ws3.append(['JobID', 'JobName', 'TaskID', 'TaskName', 'M_A', 'M_B', 'M_C'])
    rows = [
        #  JobID  Name      TaskID  Task      M_A  M_B  M_C
        [1, 'OrderA', 1, 'Cut',     5,  0,  8],
        [1, 'OrderA', 2, 'Weld',    0,  6,  9],
        [1, 'OrderA', 3, 'Finish',  4,  0,  6],

        [2, 'OrderB', 1, 'Cut',     6,  0,  9],
        [2, 'OrderB', 2, 'Weld',    0,  7, 10],
        [2, 'OrderB', 3, 'Finish',  5,  0,  7],

        [3, 'OrderC', 1, 'Cut',     4,  0,  7],
        [3, 'OrderC', 2, 'Weld',    0,  5,  8],
        [3, 'OrderC', 3, 'Finish',  3,  0,  5],

        [4, 'OrderD', 1, 'Cut',     7,  0, 10],
        [4, 'OrderD', 2, 'Weld',    0,  8, 11],
        [4, 'OrderD', 3, 'Finish',  6,  0,  8],
    ]
    for r in rows:
        ws3.append(r)

    # style headers
    BLUE_FILL = PatternFill('solid', fgColor='1A3C5E')
    for ws_s in [ws, ws2, ws3]:
        for cell in ws_s[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = BLUE_FILL
            cell.alignment = Alignment(horizontal='center')
        ws_s.column_dimensions['A'].width = 8
        for col_letter in 'BCDE':
            ws_s.column_dimensions[col_letter].width = 18

    wb.save(EXCEL_FILE)
    print(f"✓ สร้างไฟล์ตัวอย่าง: {EXCEL_FILE}")


if not os.path.exists(EXCEL_FILE):
    create_sample_excel()


# ════════════════════════════════════════════════════════════════
# STEP 1: อ่านข้อมูลจาก Excel
# ════════════════════════════════════════════════════════════════

print("=" * 60)
print("STEP 1 — อ่านข้อมูลจาก Excel")
print("=" * 60)

df_jobs = pd.read_excel(EXCEL_FILE, sheet_name='orders')
df_mach = pd.read_excel(EXCEL_FILE, sheet_name='machines')
df_proc = pd.read_excel(EXCEL_FILE, sheet_name='proc_times')

job_names  = df_jobs['JobName'].tolist()
mach_names = df_mach['MachineName'].tolist()
task_names = df_proc[df_proc['JobID'] == 1]['TaskName'].tolist()

n_job     = len(job_names)
n_task    = len(task_names)
n_machine = len(mach_names)

print(f"  Jobs     ({n_job}): {job_names}")
print(f"  Tasks    ({n_task}): {task_names}")
print(f"  Machines ({n_machine}): {mach_names}")

# สร้าง dur[j, t, m] — flatten เป็น list 1D สำหรับ MiniZinc
# ลำดับ: j วนนอก, t วนกลาง, m วนใน (ตรงกับ array3d ของ MiniZinc)
dur_flat = []
for j in range(1, n_job + 1):
    for t in range(1, n_task + 1):
        row = df_proc[(df_proc['JobID'] == j) & (df_proc['TaskID'] == t)]
        for m in mach_names:
            dur_flat.append(int(row[m].values[0]))

print(f"\n  Processing times (dur[job, task, machine]):")
idx = 0
for j, jname in enumerate(job_names):
    for t, tname in enumerate(task_names):
        vals = dur_flat[idx:idx + n_machine]
        print(f"    {jname:8s} × {tname:8s}: {dict(zip(mach_names, vals))}")
        idx += n_machine


# ════════════════════════════════════════════════════════════════
# STEP 2: รัน MiniZinc
# ════════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("STEP 2 — รัน MiniZinc (Gecode solver)")
print("=" * 60)
print(f"  Model: {MODEL_FILE}")

model    = minizinc.Model([MODEL_FILE])
solver   = minizinc.Solver.lookup("gecode")
instance = minizinc.Instance(solver, model)

# ส่ง parameters
instance["n_job"]     = n_job
instance["n_task"]    = n_task
instance["n_machine"] = n_machine
instance["dur"]       = dur_flat   # minizinc Python API รับ list 1D แล้วแปลงให้เอง

print("  กำลัง solve... ", end='', flush=True)
result = instance.solve()
print("เสร็จ!")

if not result.status.has_solution():
    print(f"  ✗ ไม่พบ solution (status: {result.status})")
    sys.exit(1)

makespan  = result["makespan"]
start_raw = result["start"]    # อาจเป็น flat list หรือ nested list ขึ้นกับ version
mach_raw  = result["machine"]

# minizinc Python API คืน 2D array เป็น nested list ใน version ใหม่
# หรือ flat list ใน version เก่า — normalize ทั้งสองแบบให้เป็น 2D [job][task]
def to_2d(raw, n_rows, n_cols):
    lst = list(raw)
    if lst and hasattr(lst[0], '__iter__'):   # nested list (API ใหม่)
        return [list(row) for row in lst]
    else:                                      # flat list (API เก่า)
        return [[lst[r * n_cols + c] for c in range(n_cols)] for r in range(n_rows)]

start   = to_2d(start_raw, n_job, n_task)
machine = to_2d(mach_raw,  n_job, n_task)

print(f"\n  ✓ Optimal makespan = {makespan} นาที")
print(f"\n  {'Job':10s} {'Task':8s} {'Machine':10s} {'Start':>6} {'End':>6} {'Dur':>5}")
print("  " + "-" * 47)
for j in range(n_job):
    for t in range(n_task):
        m_idx  = machine[j][t] - 1   # แปลงจาก 1-indexed เป็น 0-indexed
        dur_jt = dur_flat[j * n_task * n_machine + t * n_machine + m_idx]
        end    = start[j][t] + dur_jt
        print(f"  {job_names[j]:10s} {task_names[t]:8s} {mach_names[m_idx]:10s} "
              f"{start[j][t]:>6} {end:>6} {dur_jt:>5}")


# ════════════════════════════════════════════════════════════════
# STEP 3: แสดงผลด้วย matplotlib — Gantt Chart
# ════════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("STEP 3 — วาด Gantt Chart ด้วย matplotlib")
print("=" * 60)

JOB_COLORS = ['#2E86AB', '#D97706', '#1A6B47', '#DC2626',
               '#6D28D9', '#0369A1', '#92400E', '#065F46']

fig, ax = plt.subplots(figsize=(14, 5))
fig.patch.set_facecolor('#F7FAFC')
ax.set_facecolor('#FFFFFF')

legend_patches = []

for j in range(n_job):
    color = JOB_COLORS[j % len(JOB_COLORS)]
    legend_patches.append(mpatches.Patch(color=color, label=job_names[j]))

    for t in range(n_task):
        m_idx  = machine[j][t] - 1
        dur_jt = dur_flat[j * n_task * n_machine + t * n_machine + m_idx]
        s      = start[j][t]

        # วาด bar (y = machine index, x = start..end)
        bar = ax.barh(
            y         = m_idx,
            width     = dur_jt,
            left      = s,
            height    = 0.65,
            color     = color,
            edgecolor = 'white',
            linewidth = 1.5,
            alpha     = 0.9
        )

        # Label ชื่อ job + task บน bar
        if dur_jt >= 3:
            ax.text(
                s + dur_jt / 2, m_idx,
                f"{job_names[j]}\n{task_names[t]}",
                ha='center', va='center',
                fontsize=8.5, fontweight='bold', color='white'
            )

# เส้น Makespan
ax.axvline(x=makespan, color='#DC2626', linestyle='--', linewidth=2.0, zorder=5)
ax.text(makespan + 0.3, n_machine - 0.2,
        f'Makespan\n= {makespan} min',
        color='#DC2626', fontsize=9, fontweight='bold', va='top')

# แกนและ labels
ax.set_yticks(range(n_machine))
ax.set_yticklabels(mach_names, fontsize=12, fontweight='bold')
ax.set_xlabel('เวลา (นาที)', fontsize=12)
ax.set_title(
    f'Optimal Job Shop Schedule  |  Makespan = {makespan} นาที\n'
    f'(Solved by MiniZinc / Gecode  ←  ข้อมูลจาก capstone_data.xlsx)',
    fontsize=12, fontweight='bold', color='#1A3C5E', pad=10
)
ax.legend(handles=legend_patches, loc='lower right', fontsize=10,
          framealpha=0.9, title='Orders', title_fontsize=10)
ax.grid(axis='x', alpha=0.25, linestyle=':')
ax.set_xlim(0, makespan + 5)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

plt.tight_layout()
plt.savefig(OUTPUT_IMG, dpi=150, bbox_inches='tight', facecolor='#F7FAFC')
print(f"  ✓ บันทึก Gantt chart: {OUTPUT_IMG}")

plt.show()

print("\n✅ เสร็จสิ้น! ลองแก้ข้อมูลใน capstone_data.xlsx แล้วรันใหม่")
